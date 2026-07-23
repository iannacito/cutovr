#!/usr/bin/env python3
"""Local simulation of April 2021 GL through the production pipeline.

Tests whether the transfer fix (plus TotalRec) resolves the accounts 1011/1012 variance.
This does NOT post to QBO — it simulates locally using actual production functions.
"""

import sys
import openpyxl
from pathlib import Path
from decimal import Decimal
from collections import OrderedDict

sys.path.insert(0, str(Path(__file__).parent))

from pclaw_pipeline import load_general_ledger_csv, group_rows_by_transaction, money
from gl_grouping import plan_total_recoveries_group, plan_transfer_pairs, auto_balance_by_token_group

GL_DIR = Path(r"C:\Users\atala\Desktop\SampleLaw PCLawFiles\Original Files")

def load_excel_gl_openpyxl(xlsx_path):
    """Load GL directly from Excel using openpyxl."""
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active

        # Find header
        header_row = None
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row and any(row):
                row_str = str(row)
                if 'Entry Number' in row_str or 'Date' in row_str:
                    header_row = i
                    headers = [str(h).strip() if h else "" for h in row]
                    break

        if not header_row:
            return None

        # Build column map
        col_idx = {}
        for i, h in enumerate(headers):
            h_lower = h.lower()
            if 'entry' in h_lower and 'number' in h_lower:
                col_idx['transaction_id'] = i
            elif 'date' in h_lower:
                col_idx['date'] = i
            elif 'account' in h_lower and ('nickname' in h_lower or 'number' in h_lower):
                col_idx['account_number'] = i
            elif 'account' in h_lower and 'name' in h_lower:
                col_idx['account_name'] = i
            elif 'debit' in h_lower:
                col_idx['debit'] = i
            elif 'credit' in h_lower:
                col_idx['credit'] = i
            elif 'explanation' in h_lower or 'description' in h_lower:
                col_idx['description'] = i

        # Load rows
        rows = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not any(row):
                continue

            txn_id = row[col_idx.get('transaction_id')] if col_idx.get('transaction_id') else None
            if not txn_id:
                continue

            row_dict = {
                "date": str(row[col_idx['date']])[:10] if col_idx.get('date') else "",
                "account_number": str(row[col_idx['account_number']] or "").strip() if col_idx.get('account_number') else "",
                "account_name": str(row[col_idx.get('account_name')] or "").strip() if col_idx.get('account_name') else "",
                "transaction_id": str(txn_id).strip(),
                "debit": money(row[col_idx.get('debit')] if col_idx.get('debit') else 0),
                "credit": money(row[col_idx.get('credit')] if col_idx.get('credit') else 0),
                "description": str(row[col_idx.get('description')] or "").strip() if col_idx.get('description') else "",
            }
            rows.append(row_dict)

        return rows

    except Exception as e:
        print(f"Error loading {xlsx_path.name}: {e}")
        import traceback
        traceback.print_exc()
        return None

def main():
    print("="*160)
    print("LOCAL SIMULATION: Transfer Fix + TotalRec → April 2021 Accounts 1011/1012")
    print("="*160 + "\n")

    # Load April 2021 GL
    april_rows = load_excel_gl_openpyxl(GL_DIR / "GLa04888_GL - 2021-04.xlsx")
    if not april_rows:
        print("Failed to load April 2021 GL")
        return

    print(f"Loaded {len(april_rows)} rows from April 2021 GL\n")

    # ========================================================================
    # EXPECTED: What accounts 1011/1012 should net to (from source GL)
    # ========================================================================
    print("="*160)
    print("EXPECTED (from source GL)")
    print("="*160 + "\n")

    expected_1011_debits = Decimal("0")
    expected_1011_credits = Decimal("0")
    expected_1012_debits = Decimal("0")
    expected_1012_credits = Decimal("0")

    for row in april_rows:
        if row["account_number"] == "1011":
            expected_1011_debits += row["debit"]
            expected_1011_credits += row["credit"]
        elif row["account_number"] == "1012":
            expected_1012_debits += row["debit"]
            expected_1012_credits += row["credit"]

    expected_1011_net = expected_1011_debits - expected_1011_credits
    expected_1012_net = expected_1012_debits - expected_1012_credits
    expected_combined_net = expected_1011_net + expected_1012_net

    print("Account 1011 (expected from GL):")
    print(f"  Debits:  ${expected_1011_debits:>14.2f}")
    print(f"  Credits: ${expected_1011_credits:>14.2f}")
    print(f"  Net:     ${expected_1011_net:>14.2f}")
    print()
    print("Account 1012 (expected from GL):")
    print(f"  Debits:  ${expected_1012_debits:>14.2f}")
    print(f"  Credits: ${expected_1012_credits:>14.2f}")
    print(f"  Net:     ${expected_1012_net:>14.2f}")
    print()
    print(f"Combined expected variance: ${expected_combined_net:.2f}")
    print()

    # ========================================================================
    # ACTUAL: Run through the production pipeline
    # ========================================================================
    print("="*160)
    print("ACTUAL (via production pipeline)")
    print("="*160 + "\n")

    # Step 1: Group by transaction (required by TotalRec and Transfer)
    grouped = group_rows_by_transaction(april_rows)
    print(f"Grouped into {len(grouped)} transactions\n")

    # Step 2: TotalRec grouping (handles "Expense Recovery" + "Refund" patterns)
    tot_rec_groups = plan_total_recoveries_group(grouped)
    tot_rec_txn_ids = set()
    for grp in tot_rec_groups:
        tot_rec_txn_ids.update(grp.get("transaction_ids", []))
    print(f"TotalRec found {len(tot_rec_groups)} groups covering {len(tot_rec_txn_ids)} transactions\n")

    # Step 3: Transfer pairing (finds inter-account transfers)
    # NOTE: Without account_mappings/qbo_account_type_index, only keyword-based matching works
    transfer_groups = plan_transfer_pairs(grouped)
    transfer_txn_ids = set()
    for grp in transfer_groups:
        transfer_txn_ids.update(grp.get("transaction_ids", []))
    print(f"Transfer pairing found {len(transfer_groups)} pairs covering {len(transfer_txn_ids)} transactions")
    if transfer_groups:
        total_transfer_volume = sum(
            grp.get("debits", Decimal("0")) for grp in transfer_groups
        )
        print(f"Total transfer volume: ${total_transfer_volume:.2f}\n")
    else:
        print("(No transfer pairs found)\n")

    # Step 4: Auto-balance for remaining single-sided transactions
    # Detect which rows are still blocked (single-sided, not in TotalRec or Transfer groups)
    still_blocked = []
    for txn_id, txn_rows in grouped.items():
        if txn_id in tot_rec_txn_ids or txn_id in transfer_txn_ids:
            continue
        # Check if this transaction is single-sided (unbalanced-alone)
        total_debits = sum(r.get("debit", Decimal("0")) for r in txn_rows)
        total_credits = sum(r.get("credit", Decimal("0")) for r in txn_rows)
        if (total_debits > 0 and total_credits == 0) or (total_debits == 0 and total_credits > 0):
            still_blocked.append({"transaction_id": txn_id, "debits": total_debits, "credits": total_credits})

    print(f"Still-blocked (single-sided, not TotalRec/Transfer): {len(still_blocked)} transactions")
    if still_blocked:
        print("Running auto_balance_by_token_group on remaining single-sided rows...\n")
        synthetic = auto_balance_by_token_group(
            still_blocked=still_blocked,
            original_rows=april_rows,
            bank_account_name="Operating Acct.",
            bank_account_number="1012",
            expense_offset_name="Bank Charges/Fees",
            expense_offset_number="6089",
        )
        if synthetic:
            print(f"auto_balance created {len(synthetic)} synthetic rows\n")
        else:
            print("auto_balance did not create synthetics\n")
    else:
        print("No single-sided rows to auto-balance\n")
        synthetic = []

    # ========================================================================
    # COMPUTE: What 1011/1012 would be after pipeline
    # ========================================================================
    print("="*160)
    print("COMPUTED (after pipeline)")
    print("="*160 + "\n")

    # Rows to post = original + synthetic
    rows_to_post = april_rows + synthetic

    actual_1011_debits = Decimal("0")
    actual_1011_credits = Decimal("0")
    actual_1012_debits = Decimal("0")
    actual_1012_credits = Decimal("0")

    for row in rows_to_post:
        if row.get("account_number") == "1011":
            actual_1011_debits += money(row.get("debit", "0"))
            actual_1011_credits += money(row.get("credit", "0"))
        elif row.get("account_number") == "1012":
            actual_1012_debits += money(row.get("debit", "0"))
            actual_1012_credits += money(row.get("credit", "0"))

    actual_1011_net = actual_1011_debits - actual_1011_credits
    actual_1012_net = actual_1012_debits - actual_1012_credits
    actual_combined_net = actual_1011_net + actual_1012_net

    print("Account 1011 (after pipeline):")
    print(f"  Debits:  ${actual_1011_debits:>14.2f}")
    print(f"  Credits: ${actual_1011_credits:>14.2f}")
    print(f"  Net:     ${actual_1011_net:>14.2f}")
    print()
    print("Account 1012 (after pipeline):")
    print(f"  Debits:  ${actual_1012_debits:>14.2f}")
    print(f"  Credits: ${actual_1012_credits:>14.2f}")
    print(f"  Net:     ${actual_1012_net:>14.2f}")
    print()
    print(f"Combined computed variance: ${actual_combined_net:.2f}")
    print()

    # ========================================================================
    # VERDICT: Do they match?
    # ========================================================================
    print("="*160)
    print("VERDICT")
    print("="*160 + "\n")

    variance_matches = abs(expected_combined_net - actual_combined_net) < Decimal("0.01")

    print(f"Expected variance (source GL):  ${expected_combined_net:>14.2f}")
    print(f"Computed variance (after fix):  ${actual_combined_net:>14.2f}")
    print()

    if variance_matches:
        print("✓ PASS: Variances match — April 2021 would balance correctly after the fix")
    else:
        print(f"❌ FAIL: Variance mismatch of ${abs(expected_combined_net - actual_combined_net):.2f}")
        print("         The fix does not resolve the April 2021 accounts 1011/1012 issue")
    print()

if __name__ == "__main__":
    main()
