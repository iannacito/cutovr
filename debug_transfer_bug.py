#!/usr/bin/env python3
"""Debug the self-match bug and GL imbalance claims."""

import openpyxl
from pathlib import Path
from decimal import Decimal

GL_DIR = Path(r"C:\Users\atala\Desktop\SampleLaw PCLawFiles\Original Files")

def money(value):
    if value is None or value == "":
        return Decimal("0.00")
    cleaned = str(value).replace(",", "").replace("$", "").strip() or "0"
    return Decimal(cleaned).quantize(Decimal("0.01"))

def load_excel_gl(xlsx_path):
    """Load GL with detailed debugging info."""
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active

        # Find header
        header_row = None
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row and any(row):
                row_str = str(row)
                if 'Entry Number' in row_str or 'Date' in row_str:
                    header_row = i
                    headers = [str(h).strip() if h else "" for h in row]
                    break

        if not header_row:
            return None, None

        # Build column map
        col_idx = {}
        for i, h in enumerate(headers):
            h_lower = h.lower()
            if 'entry' in h_lower and 'number' in h_lower:
                col_idx['transaction_id'] = i
            elif 'date' in h_lower:
                col_idx['date'] = i
            elif 'account' in h_lower and ('nickname' in h_lower or 'number' in h_lower):
                col_idx['account_number'] = i
            elif 'account' in h_lower and 'name' in h_lower:
                col_idx['account_name'] = i
            elif 'debit' in h_lower:
                col_idx['debit'] = i
            elif 'credit' in h_lower:
                col_idx['credit'] = i
            elif 'explanation' in h_lower or 'description' in h_lower:
                col_idx['description'] = i

        # Load rows with full detail
        rows = []
        for row_num, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
            if not row or not any(row):
                continue

            txn_id = row[col_idx.get('transaction_id')] if col_idx.get('transaction_id') else None
            if not txn_id:
                continue

            row_dict = {
                "row_num": row_num,
                "date": str(row[col_idx['date']])[:10] if col_idx.get('date') else "",
                "account_number": str(row[col_idx['account_number']] or "").strip() if col_idx.get('account_number') else "",
                "account_name": str(row[col_idx.get('account_name')] or "").strip() if col_idx.get('account_name') else "",
                "transaction_id": str(txn_id).strip(),
                "debit": money(row[col_idx.get('debit')] if col_idx.get('debit') else 0),
                "credit": money(row[col_idx.get('credit')] if col_idx.get('credit') else 0),
                "description": str(row[col_idx.get('description')] or "").strip() if col_idx.get('description') else "",
            }
            rows.append(row_dict)

        return rows, headers

    except Exception as e:
        print(f"Error loading {xlsx_path.name}: {e}")
        return None, None

def main():
    print("="*160)
    print("DEBUG: Entry 261504 Self-Match Bug")
    print("="*160 + "\n")

    april_rows, headers = load_excel_gl(GL_DIR / "GLa04888_GL - 2021-04.xlsx")
    if not april_rows:
        print("Failed to load April 2021 GL")
        return

    # Find all rows with transaction_id 261504
    entry_261504 = [r for r in april_rows if r["transaction_id"] == "261504"]
    print(f"Rows with entry 261504: {len(entry_261504)}")
    for row in entry_261504:
        print(f"  Row {row['row_num']}: Acct {row['account_number']} "
              f"Debit=${row['debit']:.2f} Credit=${row['credit']:.2f} "
              f"— {row['description'][:60]}")
    print()

    # Show entries around 261504 in transaction_id order
    print("Entries 261500–261510 (to check for duplicates/related entries):")
    for i in range(261500, 261511):
        entry_rows = [r for r in april_rows if r["transaction_id"] == str(i)]
        if entry_rows:
            for row in entry_rows:
                print(f"  Entry {row['transaction_id']:>6} (row {row['row_num']:>4}): "
                      f"Acct {row['account_number']:>4} Debit=${row['debit']:>10.2f} "
                      f"Credit=${row['credit']:>10.2f} — {row['description'][:50]}")
    print()

    # ========================================================================
    # Now debug the GL imbalance claim
    # ========================================================================
    print("="*160)
    print("DEBUG: 'Every Month Imbalanced' Claim")
    print("="*160 + "\n")

    # Check April specifically
    total_debits_april = sum(r["debit"] for r in april_rows)
    total_credits_april = sum(r["credit"] for r in april_rows)

    print(f"April 2021 GL totals (from {len(april_rows)} rows):")
    print(f"  Total debits:  ${total_debits_april:.2f}")
    print(f"  Total credits: ${total_credits_april:.2f}")
    print(f"  Difference:    ${abs(total_debits_april - total_credits_april):.2f}")

    if abs(total_debits_april - total_credits_april) < Decimal("0.01"):
        print(f"  ✓ BALANCED")
    else:
        print(f"  ❌ IMBALANCED — this contradicts known fact 'Jan-Mar 2021: Balanced ✓'")

    print()
    print("Checking row-level structure...")

    # Count rows by their debit/credit pattern
    both_sided = [r for r in april_rows if r["debit"] > 0 and r["credit"] > 0]
    debit_only = [r for r in april_rows if r["debit"] > 0 and r["credit"] == 0]
    credit_only = [r for r in april_rows if r["debit"] == 0 and r["credit"] > 0]
    neither = [r for r in april_rows if r["debit"] == 0 and r["credit"] == 0]

    print(f"  Both-sided rows (debit AND credit): {len(both_sided)}")
    print(f"  Debit-only rows: {len(debit_only)}")
    print(f"  Credit-only rows: {len(credit_only)}")
    print(f"  Neither (zero amounts): {len(neither)}")
    print()

    if both_sided:
        print("  ⚠️  Both-sided rows exist! These are unusual for GL entries.")
        print("  Showing first 5:")
        for row in both_sided[:5]:
            print(f"    Entry {row['transaction_id']} (row {row['row_num']}): "
                  f"Debit=${row['debit']:.2f} Credit=${row['credit']:.2f}")
        print()

    # ========================================================================
    # Check April 2021 accounts 1011/1012 with row-by-row detail
    # ========================================================================
    print("="*160)
    print("Accounts 1011/1012 in April 2021 (Row-by-Row Detail)")
    print("="*160 + "\n")

    acct_1011 = [r for r in april_rows if r["account_number"] == "1011"]
    acct_1012 = [r for r in april_rows if r["account_number"] == "1012"]

    print(f"Account 1011 ({len(acct_1011)} rows):")
    acct_1011_debits = Decimal("0")
    acct_1011_credits = Decimal("0")
    for row in acct_1011:
        acct_1011_debits += row["debit"]
        acct_1011_credits += row["credit"]
        if row["debit"] > 0 or row["credit"] > 0:
            print(f"  Entry {row['transaction_id']:>6} (row {row['row_num']:>4}): "
                  f"Debit=${row['debit']:>10.2f} Credit=${row['credit']:>10.2f}")
    print(f"  Totals: Debits=${acct_1011_debits:.2f}, Credits=${acct_1011_credits:.2f}")
    print()

    print(f"Account 1012 ({len(acct_1012)} rows):")
    acct_1012_debits = Decimal("0")
    acct_1012_credits = Decimal("0")
    for row in acct_1012:
        acct_1012_debits += row["debit"]
        acct_1012_credits += row["credit"]
        if row["debit"] > 0 or row["credit"] > 0:
            print(f"  Entry {row['transaction_id']:>6} (row {row['row_num']:>4}): "
                  f"Debit=${row['debit']:>10.2f} Credit=${row['credit']:>10.2f}")
    print(f"  Totals: Debits=${acct_1012_debits:.2f}, Credits=${acct_1012_credits:.2f}")
    print()

    print(f"Combined 1011/1012:")
    print(f"  Combined debits:  ${acct_1011_debits + acct_1012_debits:.2f}")
    print(f"  Combined credits: ${acct_1011_credits + acct_1012_credits:.2f}")
    print(f"  Net: ${(acct_1011_debits - acct_1011_credits) + (acct_1012_debits - acct_1012_credits):.2f}")

if __name__ == "__main__":
    main()
